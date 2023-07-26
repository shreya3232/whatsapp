from openai import Image
import openpyxl
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
from PIL import Image as PILImage
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage

# Load the existing Excel file
input_excel_file = "Helium_10_Xray_2023-07-25 (1).xlsx"
wb = load_workbook(input_excel_file)
ws = wb.active

# Find the last column with data in the input Excel file
last_column = len(ws[1])

# Step 1: Insert the downloaded images into the new column
for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if row_num == 1:  # Skip the header row
        ws.cell(row=1, column=last_column + 1, value="Images")
        continue

    url = row[3]  # Assuming the URLs are in the first column (column A)

    response = requests.get(url)
    if response.status_code == 200:
        img_data = BytesIO(response.content)
        img = PILImage.open(img_data)
        img.thumbnail((40, 40))  # Resize the image to 40px x 40px

        cell = ws.cell(row=row_num, column=last_column + 1)
        cell.value = ""
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(last_column + 1)].width = 10
        ws.row_dimensions[row_num].height = 40
        xl_img = XLImage(img)
        xl_img.width = xl_img.height = 40
        anchor = cell.coordinate  # Get the cell's coordinate (e.g., "B2")
        ws.add_image(xl_img, anchor)

# Save the modified Excel file
wb.save(input_excel_file)

print("Images inserted into the input Excel file successfully.")
