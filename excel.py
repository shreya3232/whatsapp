import requests
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# Load the existing Excel file
input_excel_file = "Helium_10_Xray_2023-07-25 (1).xlsx"
wb = load_workbook(input_excel_file)
ws = wb.active

# Create a new Excel file to store the downloaded images
output_excel_file = "output.xlsx"
wb_output = Workbook()
ws_output = wb_output.active

# Step 1: Get the URLs from the input Excel file and download the images
urls = []
for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if row_num == 1:
        continue  # Skip the header row

    url = row[3]  # Assuming the URLs are in the first column (column A)
    urls.append(url)

    response = requests.get(url)
    if response.status_code == 200:
        with open(f"image{row_num}.jpg", "wb") as f:
            f.write(response.content)

# Step 2: Add the URLs to the output Excel file (optional)
for i, url in enumerate(urls, start=1):
    ws_output.cell(row=i, column=1, value=url)

# Step 3: Insert the downloaded images into the new Excel file
for row_num, _ in enumerate(ws_output.iter_rows(min_row=1, values_only=True), start=1):
    if row_num == 1:
        continue  # Skip the header row

    img = Image(f"image{row_num}.jpg")
    img.width = img.height = 100  # Resize the image to 40px x 40px
    cell = ws_output.cell(row=row_num, column=2)
    cell.value = ""
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    ws_output.add_image(img, f'B{row_num}')  # Use the coordinate string directly

# Save the output Excel file
wb_output.save(output_excel_file)

print("Images downloaded and inserted into Excel file successfully.")
